

import json
import http.client
from urllib.parse import urlencode
import asyncio
import argparse
import os
import re
import random
from datetime import datetime
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeout
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import requests
from dotenv import load_dotenv


# ---------------- Configuration ----------------
load_dotenv()
SCRAPER_API_KEY = os.getenv("key")
SCOPUS_API_KEY = os.getenv("scopus_key")

if not SCRAPER_API_KEY:
    raise SystemExit("âŒ No ScraperAPI key found in .env file. Add 'key=your_api_key'")

# User agents to rotate (anti-detection)
USER_AGENTS = [
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:121.0) Gecko/20100101 Firefox/121.0',
]


# ---------------- Scopus API Functions ----------------
def get_scopus_author_id(author_name):
    """Search for author in Scopus and return their Scopus ID"""
    if not SCOPUS_API_KEY:
        print("âš ï¸  No Scopus API key found. Skipping Scopus data.")
        return None
    
    try:
        url = "https://api.elsevier.com/content/search/author"
        headers = {
            "X-ELS-APIKey": SCOPUS_API_KEY,
            "Accept": "application/json"
        }
        params = {
            "query": f"AUTHLAST({author_name.split()[-1]}) AND AUTHFIRST({author_name.split()[0]})",
            "count": 5
        }
        
        response = requests.get(url, headers=headers, params=params, timeout=10)
        
        if response.status_code == 200:
            data = response.json()
            results = data.get("search-results", {}).get("entry", [])
            
            if results and len(results) > 0:
                author_id = results[0].get("dc:identifier", "").replace("AUTHOR_ID:", "")
                author_full_name = results[0].get("preferred-name", {}).get("given-name", "") + " " + \
                                 results[0].get("preferred-name", {}).get("surname", "")
                print(f"âœ… Found Scopus Author: {author_full_name} (ID: {author_id})")
                return author_id
        else:
            print(f"âš ï¸  Scopus API returned status {response.status_code}")
            
    except Exception as e:
        print(f"âš ï¸  Error fetching Scopus author ID: {e}")
    
    return None


def get_scopus_publication_details(author_id, publication_title):
    """Get publication details from Scopus by author ID and title"""
    if not SCOPUS_API_KEY or not author_id:
        return None
    
    try:
        url = "https://api.elsevier.com/content/search/scopus"
        headers = {
            "X-ELS-APIKey": SCOPUS_API_KEY,
            "Accept": "application/json"
        }
        
        clean_title = re.sub(r'[^\w\s]', '', publication_title).strip()
        title_words = clean_title.split()[:5]
        
        params = {
            "query": f"AU-ID({author_id}) AND TITLE({' '.join(title_words)})",
            "count": 1
        }
        
        response = requests.get(url, headers=headers, params=params, timeout=10)
        
        if response.status_code == 200:
            data = response.json()
            results = data.get("search-results", {}).get("entry", [])
            
            if results and len(results) > 0:
                pub = results[0]
                return {
                    "scopus_id": pub.get("dc:identifier", "").replace("SCOPUS_ID:", ""),
                    "eid": pub.get("eid", "N/A"),
                    "doi": pub.get("prism:doi", "N/A"),
                    "publication_year": pub.get("prism:coverDate", "N/A")[:4],
                }
                
    except Exception as e:
        print(f"  âš ï¸  Error fetching Scopus details: {e}")
    
    return None


# ---------------- Google Scholar Scraping Functions ----------------
async def scrape_google_scholar_playwright(author_name_or_url):
    """Main function to scrape Google Scholar using Playwright"""
    print(f"\n{'='*60}")
    print(f"ðŸ” Starting scrape for: {author_name_or_url}")
    print(f"{'='*60}\n")
    
    # Check if input is a URL or author name
    if author_name_or_url.startswith("http"):
        author_link = author_name_or_url
        print(f"âœ… Using direct profile URL: {author_link}\n")
        # Extract author name from URL or use generic
        author_name = "Scholar_Author"
    else:
        # Step 1: Search for author profile
        print("ðŸ“¡ Step 1: Searching for author on Google Scholar...")
        author_link = await search_author_with_playwright(author_name_or_url)
        
        if not author_link:
            raise SystemExit("âŒ Could not find author profile link.")
        
        print(f"âœ… Found author profile: {author_link}\n")
        author_name = author_name_or_url
    
    # Step 2: Get Scopus Author ID (COMMENTED OUT)
    scopus_author_id = None  # Disabled Scopus
    
    # Step 3: Scrape publications with Playwright
    print("ðŸŒ Step 3: Scraping publications with Playwright...")
    publications = await scrape_publications_with_playwright(author_link)
    
    # Step 5: Save to Excel
    print(f"\nðŸ’¾ Step 5: Saving to Excel...")
    filename = save_to_excel(publications, author_name, scopus_author_id)
    
    print(f"\n{'='*60}")
    print(f"âœ… SUCCESS! Scraped {len(publications)} publications")
    print(f"ðŸ“ Saved to: {filename}")
    print(f"{'='*60}\n")
    
    return publications


async def search_author_with_playwright(author_name):
    """Search for author on Google Scholar using Playwright with anti-detection"""
    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=False,
            args=[
                '--disable-blink-features=AutomationControlled',
                '--disable-dev-shm-usage',
                '--no-sandbox'
            ]
        )
        context = await browser.new_context(
            user_agent=random.choice(USER_AGENTS),
            viewport={'width': 1920, 'height': 1080},
            locale='en-US'
        )
        await context.add_init_script("""Object.defineProperty(navigator, 'webdriver', {get: () => undefined})""")
        
        page = await context.new_page()
        
        try:
            # Navigate to Google Scholar
            await page.goto("https://scholar.google.com", wait_until="domcontentloaded")
            await asyncio.sleep(random.uniform(2, 4))
            
            # Simulate mouse movement
            await page.mouse.move(100, 100)
            await asyncio.sleep(random.uniform(0.5, 1))
            
            # Search with human-like typing
            search_box = page.locator('input[name="q"]')
            await search_box.click()
            await asyncio.sleep(random.uniform(0.3, 0.7))
            
            for char in author_name:
                await search_box.type(char, delay=random.uniform(50, 150))
            
            await asyncio.sleep(random.uniform(0.5, 1))
            await search_box.press("Enter")
            await asyncio.sleep(random.uniform(3, 5))
            
            # Look for author profile link
            profile_links = await page.locator('a[href*="/citations?user="]').all()
            
            if profile_links and len(profile_links) > 0:
                href = await profile_links[0].get_attribute("href")
                if href:
                    author_link = href if href.startswith("http") else "https://scholar.google.com" + href
                    await browser.close()
                    return author_link
            
            # Alternative: try author name links
            author_name_links = await page.locator('.gs_ai_name a').all()
            if author_name_links and len(author_name_links) > 0:
                href = await author_name_links[0].get_attribute("href")
                if href and "/citations?user=" in href:
                    author_link = href if href.startswith("http") else "https://scholar.google.com" + href
                    await browser.close()
                    return author_link
            
            await browser.close()
            return None
            
        except Exception as e:
            print(f"âš ï¸  Error searching for author: {e}")
            await browser.close()
            return None


async def scrape_publications_with_playwright(author_link):
    """Scrape all publications from Google Scholar profile with anti-CAPTCHA measures"""
    publications = []
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=False,
            args=[
                '--disable-blink-features=AutomationControlled',
                '--disable-dev-shm-usage',
                '--no-sandbox'
            ]
        )
        context = await browser.new_context(
            user_agent=random.choice(USER_AGENTS),
            viewport={'width': 1920, 'height': 1080},
            locale='en-US'
        )
        await context.add_init_script("""Object.defineProperty(navigator, 'webdriver', {get: () => undefined})""")
        
        page = await context.new_page()
        
        try:
            # Navigate to author profile
            await page.goto(author_link, wait_until="domcontentloaded")
            await asyncio.sleep(random.uniform(2, 4))
            
            # Simulate human behavior
            await page.mouse.move(random.randint(100, 300), random.randint(100, 300))
            
            # Wait for publications table
            await page.wait_for_selector(".gsc_a_at", timeout=10000)
            
            # Click "Show more" with human-like behavior
            click_count = 0
            while True:
                try:
                    # Gradual scrolling
                    for i in range(3):
                        await page.evaluate(f"window.scrollBy(0, {random.randint(200, 400)})")
                        await asyncio.sleep(random.uniform(0.3, 0.7))
                    
                    await asyncio.sleep(random.uniform(1, 2))
                    
                    show_more = page.locator("button:has-text('Show more')")
                    
                    if await show_more.count() > 0 and await show_more.is_enabled():
                        await show_more.hover()
                        await asyncio.sleep(random.uniform(0.3, 0.6))
                        await show_more.click()
                        click_count += 1
                        print(f"  ðŸ“„ Clicked 'Show more' ({click_count} times)...")
                        await asyncio.sleep(random.uniform(2, 4))
                    else:
                        break
                        
                except PlaywrightTimeout:
                    break
                except Exception:
                    print(f"  âš ï¸  No more publications to load.")
                    break
            
            print(f"âœ… All publications loaded.\n")
            
            # Get all publication elements
            pub_elements = await page.locator(".gsc_a_tr").all()
            print(f"ðŸ“š Found {len(pub_elements)} publications. Extracting details...\n")
            
            # Extract publication data
            for i, pub_row in enumerate(pub_elements, 1):
                try:
                    title_elem = pub_row.locator(".gsc_a_at")
                    title = await title_elem.inner_text()
                    href = await title_elem.get_attribute("href")
                    
                    # Get year and citations
                    year_elem = pub_row.locator(".gsc_a_y span")
                    year = await year_elem.inner_text() if await year_elem.count() > 0 else "N/A"
                    
                    citations_elem = pub_row.locator(".gsc_a_c a").first
                    citations = "0"
                    try:
                        citations = await citations_elem.inner_text() if await citations_elem.count() > 0 else "0"
                        citations = citations if citations else "0"
                    except:
                        pass
                    
                    print(f"{i}. {title[:60]}... (Year: {year}, Citations: {citations})")
                    
                    # Add longer delay between requests (5-8 seconds to avoid rate limiting)
                    print(f"    â³ Waiting {random.randint(5, 8)} seconds before next request...")
                    await asyncio.sleep(random.uniform(5, 8))
                    
                    # Get abstract
                    abstract = await get_abstract_from_publication_page(browser, href, i)
                    
                    publications.append({
                        "title": title.strip(),
                        "year": year,
                        "citations": citations,
                        "abstract": abstract,
                        "scopus_id": None,
                        "scopus_eid": None,
                        "scopus_doi": None,
                        "scopus_year": None
                    })
                    
                except Exception as e:
                    print(f"  âš ï¸  Error extracting publication {i}: {e}")
                    continue
            
        finally:
            await browser.close()
    
    return publications


async def get_abstract_from_publication_page(browser, pub_href, pub_num):
    """Navigate to publication page and extract abstract with aggressive anti-CAPTCHA measures"""
    if not pub_href:
        return "(No abstract found)"
    
    abstract_text = "(No abstract found)"
    new_page = None
    context = None
    
    try:
        # Create fresh context with new user agent and more realistic settings
        context = await browser.new_context(
            user_agent=random.choice(USER_AGENTS),
            viewport={'width': 1920, 'height': 1080},
            locale='en-US',
            timezone_id='America/New_York',
            geolocation={'longitude': -74.0060, 'latitude': 40.7128},
            permissions=['geolocation']
        )
        
        # Advanced stealth
        await context.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
            Object.defineProperty(navigator, 'plugins', {get: () => [1, 2, 3, 4, 5]});
            Object.defineProperty(navigator, 'languages', {get: () => ['en-US', 'en']});
            window.chrome = {runtime: {}};
        """)
        
        new_page = await context.new_page()
        
        # Add extra headers to look more realistic
        await new_page.set_extra_http_headers({
            'Accept-Language': 'en-US,en;q=0.9',
            'Accept-Encoding': 'gzip, deflate, br',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'DNT': '1',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1'
        })
        
        full_url = f"https://scholar.google.com{pub_href}" if pub_href.startswith("/") else pub_href
        
        # Navigate with extended timeout
        await new_page.goto(full_url, wait_until="domcontentloaded", timeout=20000)
        await asyncio.sleep(random.uniform(3, 5))  # Longer initial wait
        
        # Check for CAPTCHA
        captcha_present = await new_page.locator("#gs_captcha_f").count() > 0
        if captcha_present:
            print(f"    ðŸš« CAPTCHA detected! Manual intervention may be needed.")
            print(f"    â¸ï¸  Waiting 30 seconds for manual solve...")
            await asyncio.sleep(30)  # Give time to manually solve
            
            # Check again after wait
            captcha_still_present = await new_page.locator("#gs_captcha_f").count() > 0
            if captcha_still_present:
                await new_page.close()
                await context.close()
                return "(CAPTCHA blocked - solve manually)"
        
        # More realistic human reading behavior
        # Random mouse movements
        for _ in range(3):
            await new_page.mouse.move(
                random.randint(200, 800),
                random.randint(200, 600)
            )
            await asyncio.sleep(random.uniform(0.3, 0.7))
        
        # Gradual scrolling like a human reading
        for _ in range(4):
            await new_page.evaluate(f"window.scrollBy(0, {random.randint(100, 250)})")
            await asyncio.sleep(random.uniform(0.8, 1.5))
        
        # Try expanded list of selectors for abstract
        selectors = [
            ".gsh_csp",                  # Main abstract container
            ".gsh_csp_ab",               # Abstract text
            ".gsh_small",                # Small text
            "#gsc_oci_merged",           # Merged info section
            ".gs_scl",                   # Scholar content
            "div.gsh_csp",               # Div with class
            "div[style*='text-align']",  # Generic text divs
            ".gsc_vcd_value",            # Value container
            ".gsc_oci_value",            # OCI value
            "div#gsc_vcd_table div",     # Table divs
            ".gs_rs",                    # Result snippet
        ]
        
        for selector in selectors:
            try:
                elem = new_page.locator(selector).first
                if await elem.count() > 0:
                    text = await elem.inner_text(timeout=5000)
                    if text and len(text.strip()) > 20:
                        abstract_text = text.strip()
                        print(f"    âœ… Abstract extracted ({len(abstract_text)} chars)")
                        break
            except:
                continue
        
        # If still no abstract, try getting all text from the page
        if abstract_text == "(No abstract found)":
            try:
                # Look for any div that might contain abstract-like text
                all_divs = await new_page.locator("div").all()
                for div in all_divs[:20]:  # Check first 20 divs
                    try:
                        text = await div.inner_text(timeout=2000)
                        # Check if it looks like an abstract (long text, not navigation)
                        if text and 100 < len(text.strip()) < 3000 and not any(word in text.lower() for word in ['citation', 'export', 'copyright', 'menu']):
                            abstract_text = text.strip()
                            print(f"    âœ… Abstract found via scanning ({len(abstract_text)} chars)")
                            break
                    except:
                        continue
            except:
                pass
        
        # Small delay before closing
        await asyncio.sleep(random.uniform(1, 2))
        await new_page.close()
        await context.close()
        
    except Exception as e:
        print(f"    âš ï¸  Error: {str(e)[:100]}")
        if new_page:
            try:
                await new_page.close()
            except:
                pass
        if context:
            try:
                await context.close()
            except:
                pass
    
    return abstract_text


def enrich_with_scopus_data(publications, scopus_author_id):
    """Enrich publications with Scopus data"""
    for i, pub in enumerate(publications, 1):
        print(f"  {i}/{len(publications)} Checking Scopus for: {pub['title'][:50]}...")
        
        scopus_data = get_scopus_publication_details(scopus_author_id, pub["title"])
        
        if scopus_data:
            pub["scopus_id"] = scopus_data.get("scopus_id", "N/A")
            pub["scopus_eid"] = scopus_data.get("eid", "N/A")
            pub["scopus_doi"] = scopus_data.get("doi", "N/A")
            pub["scopus_year"] = scopus_data.get("publication_year", "N/A")
            print(f"    âœ… Found in Scopus (Year: {pub['scopus_year']})")
        else:
            print(f"    âš ï¸  Not found in Scopus")
    
    return publications


def save_to_excel(publications, author_name, scopus_author_id):
    """Save publications to Excel file with formatting"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Publications"
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Headers
    headers = [
        "No.",
        "Author Name",
        "Scopus Author ID",
        "Publication Title",
        "Abstract",
        "Publication Year (Scholar)",
        "Publication Year (Scopus)",
        "Citations (Scholar)",
        "Scopus Document ID",
        "Scopus EID",
        "DOI"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 80
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 30
    ws.column_dimensions['K'].width = 25
    
    # Add data
    for i, pub in enumerate(publications, 1):
        row = i + 1
        ws.cell(row, 1, i)
        ws.cell(row, 2, author_name)
        ws.cell(row, 3, scopus_author_id or "N/A")
        ws.cell(row, 4, pub["title"])
        ws.cell(row, 5, pub["abstract"])
        ws.cell(row, 6, pub["year"])
        ws.cell(row, 7, pub.get("scopus_year", "N/A"))
        ws.cell(row, 8, pub["citations"])
        ws.cell(row, 9, pub.get("scopus_id", "N/A"))
        ws.cell(row, 10, pub.get("scopus_eid", "N/A"))
        ws.cell(row, 11, pub.get("scopus_doi", "N/A"))
        
        # Apply alignment
        for col in range(1, 12):
            ws.cell(row, col).alignment = Alignment(vertical="top", wrap_text=True)
    
    # Save file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"publications_{author_name.replace(' ', '_')}_{timestamp}.xlsx"
    wb.save(filename)
    
    return filename


# ---------------- Main Entry Point ----------------
async def main():
    parser = argparse.ArgumentParser(
        description="Scrape publications from Google Scholar and enrich with Scopus data"
    )
    parser.add_argument(
        "author_name",
        type=str,
        nargs="?",
        help="Author name to search for (e.g., 'John Smith')"
    )
    
    args = parser.parse_args()
    
    if not args.author_name:
        author_name = input("Enter author name: ").strip()
        if not author_name:
            raise SystemExit("âŒ Author name is required.")
    else:
        author_name = args.author_name
    
    await scrape_google_scholar_playwright(author_name)


if __name__ == "__main__":
    asyncio.run(main())
