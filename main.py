
import json
import http.client
from urllib.parse import urlencode
import asyncio
import argparse
import os
import re
from datetime import datetime
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeout
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import requests
from dotenv import load_dotenv



load_dotenv()
SCRAPER_API_KEY = os.getenv("key")
SCOPUS_API_KEY = os.getenv("scopus_key")  

if not SCRAPER_API_KEY:
    raise SystemExit("No ScraperAPI key found in .env file. Add 'key=your_api_key'")



def get_scopus_author_id(author_name):
    """Search for author in Scopus and return their Scopus ID"""
    if not SCOPUS_API_KEY:
        print("No Scopus API key found. Skipping Scopus data.")
        return None
    
    try:
        url = "https://api.elsevier.com/content/search/author"
        headers = {
            "X-ELS-APIKey": SCOPUS_API_KEY,
            "Accept": "application/json"
        }
        params = {
            "query": f"AUTHLAST({author_name.split()[-1]}) AND AUTHFIRST({author_name.split()[0]})",
            "count": 5
        }
        
        response = requests.get(url, headers=headers, params=params, timeout=10)
        
        if response.status_code == 200:
            data = response.json()
            results = data.get("search-results", {}).get("entry", [])
            
            if results and len(results) > 0:
              
                author_id = results[0].get("dc:identifier", "").replace("AUTHOR_ID:", "")
                author_full_name = results[0].get("preferred-name", {}).get("given-name", "") + " " + \
                                 results[0].get("preferred-name", {}).get("surname", "")
                print(f"‚úÖ Found Scopus Author: {author_full_name} (ID: {author_id})")
                return author_id
        else:
            print(f"‚ö†Ô∏è  Scopus API returned status {response.status_code}")
            
    except Exception as e:
        print(f"‚ö†Ô∏è  Error fetching Scopus author ID: {e}")
    
    return None


def get_scopus_publication_details(author_id, publication_title):
    """Get publication details from Scopus by author ID and title"""
    if not SCOPUS_API_KEY or not author_id:
        return None
    
    try:
        url = "https://api.elsevier.com/content/search/scopus"
        headers = {
            "X-ELS-APIKey": SCOPUS_API_KEY,
            "Accept": "application/json"
        }
        
        
        clean_title = re.sub(r'[^\w\s]', '', publication_title).strip()
        title_words = clean_title.split()[:5] 
        
        params = {
            "query": f"AU-ID({author_id}) AND TITLE({' '.join(title_words)})",
            "count": 1
        }
        
        response = requests.get(url, headers=headers, params=params, timeout=10)
        
        if response.status_code == 200:
            data = response.json()
            results = data.get("search-results", {}).get("entry", [])
            
            if results and len(results) > 0:
                pub = results[0]
                return {
                    "scopus_id": pub.get("dc:identifier", "").replace("SCOPUS_ID:", ""),
                    "eid": pub.get("eid", "N/A"),
                    "doi": pub.get("prism:doi", "N/A"),
                    "publication_year": pub.get("prism:coverDate", "N/A")[:4],
                }
                
    except Exception as e:
        print(f"  ‚ö†Ô∏è  Error fetching Scopus details: {e}")
    
    return None



async def scrape_google_scholar_playwright(author_name):
    """
    Main function to scrape Google Scholar using Playwright
    """
    print(f"\n{'='*60}")
    print(f"üîç Starting scrape for author: {author_name}")
    print(f"{'='*60}\n")
    
   
    print("üì° Step 1: Searching for author on Google Scholar...")
    author_link = await search_author_with_playwright(author_name)
    
    if not author_link:
        raise SystemExit("‚ùå Could not find author profile link.")
    
    print(f"‚úÖ Found author profile: {author_link}\n")
    

    scopus_author_id = None 
    

    print("üåê Step 3: Scraping publications with Playwright...")
    publications = await scrape_publications_with_playwright(author_link)
    

    print(f"\nüíæ Step 5: Saving to Excel...")
    filename = save_to_excel(publications, author_name, scopus_author_id)
    
    print(f"\n{'='*60}")
    print(f"‚úÖ SUCCESS! Scraped {len(publications)} publications")
    print(f"üìÅ Saved to: {filename}")
    print(f"{'='*60}\n")
    
    return publications


async def search_author_with_playwright(author_name):
    """Search for author on Google Scholar using Playwright and return profile link"""
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        page = await browser.new_page()
        
        try:
           
            await page.goto("https://scholar.google.com", wait_until="domcontentloaded")
            await asyncio.sleep(2)
            
           
            search_box = page.locator('input[name="q"]')
            await search_box.fill(author_name)
            await search_box.press("Enter")
            await asyncio.sleep(3)
            

            profile_links = await page.locator('a[href*="/citations?user="]').all()
            
            if profile_links and len(profile_links) > 0:
                
                href = await profile_links[0].get_attribute("href")
                if href:
                    if href.startswith("http"):
                        author_link = href
                    else:
                        author_link = "https://scholar.google.com" + href
                    
                    await browser.close()
                    return author_link
            
            
            author_name_links = await page.locator('.gs_ai_name a').all()
            if author_name_links and len(author_name_links) > 0:
                href = await author_name_links[0].get_attribute("href")
                if href and "/citations?user=" in href:
                    if href.startswith("http"):
                        author_link = href
                    else:
                        author_link = "https://scholar.google.com" + href
                    
                    await browser.close()
                    return author_link
            
            await browser.close()
            return None
            
        except Exception as e:
            print(f"‚ö†Ô∏è  Error searching for author: {e}")
            await browser.close()
            return None


async def get_author_profile_link(author_name):
    """Use ScraperAPI to get author profile link from Google Scholar"""
    try:
        conn = http.client.HTTPSConnection("scraperapi.thordata.com")
        params = {
            "engine": "google_scholar",
            "q": author_name,
            "json": "1",
            "start": "0"
        }
        payload = urlencode(params)
        headers = {
            'Authorization': f'Bearer {SCRAPER_API_KEY}',
            'Content-Type': 'application/x-www-form-urlencoded'
        }
        
        conn.request("POST", "/request", payload, headers)
        res = conn.getresponse()
        raw = res.read().decode("utf-8")
        
        data = json.loads(raw)
        if isinstance(data, str):
            data = json.loads(data)
        
        organic = data.get("organic_results", [])
        if not organic:
            return None
        
       
        for result in organic:
            authors = result.get("publication_info", {}).get("authors", [])
            if authors:
                author_link = authors[0].get("link")
                if author_link:
                    return "https://scholar.google.com" + author_link
        
        return None
        
    except Exception as e:
        print(f"‚ùå Error fetching author profile: {e}")
        return None


async def scrape_publications_with_playwright(author_link):
    """Scrape all publications from Google Scholar profile using Playwright"""
    publications = []
    
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        page = await browser.new_page()
        
        try:
            # Navigate to author profile
            await page.goto(author_link, wait_until="domcontentloaded")
            
            # Wait for publications table
            await page.wait_for_selector(".gsc_a_at", timeout=10000)
            
            # Click "Show more" button until all publications are loaded
            click_count = 0
            while True:
                try:
                    # Scroll to bottom
                    await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                    await asyncio.sleep(1)
                    
                    # Try to find and click "Show more" button
                    show_more = page.locator("button:has-text('Show more')")
                    
                    if await show_more.count() > 0 and await show_more.is_enabled():
                        await show_more.click()
                        click_count += 1
                        print(f"  üìÑ Clicked 'Show more' ({click_count} times)...")
                        await asyncio.sleep(2)
                    else:
                        break
                        
                except PlaywrightTimeout:
                    break
                except Exception as e:
                    print(f"  ‚ö†Ô∏è  No more publications to load.")
                    break
            
            print(f"‚úÖ All publications loaded.\n")
            
            # Get all publication links - need to do this before browser closes
            pub_elements = await page.locator(".gsc_a_tr").all()
            print(f"üìö Found {len(pub_elements)} publications. Extracting details...\n")
            
            # Extract publication data
            for i, pub_row in enumerate(pub_elements, 1):
                try:
                    title_elem = pub_row.locator(".gsc_a_at")
                    title = await title_elem.inner_text()
                    href = await title_elem.get_attribute("href")
                    
                    # Get year and citations from the row
                    year_elem = pub_row.locator(".gsc_a_y span")
                    year = await year_elem.inner_text() if await year_elem.count() > 0 else "N/A"
                    
                    citations_elem = pub_row.locator(".gsc_a_c a")
                    citations = await citations_elem.inner_text() if await citations_elem.count() > 0 else "0"
                    citations = citations if citations else "0"
                    
                    print(f"{i}. {title[:60]}... (Year: {year}, Citations: {citations})")
                    
                    # Get abstract by opening publication page
                    abstract = await get_abstract_from_publication_page(browser, href)
                    
                    publications.append({
                        "title": title.strip(),
                        "year": year,
                        "citations": citations,
                        "abstract": abstract,
                        "scopus_id": None,
                        "scopus_eid": None,
                        "scopus_doi": None,
                        "scopus_year": None
                    })
                    
                except Exception as e:
                    print(f"  ‚ö†Ô∏è  Error extracting publication {i}: {e}")
                    continue
            
        finally:
            await browser.close()
    
    return publications


async def get_abstract_from_publication_page(browser, pub_href):
    """Navigate to publication page and extract abstract"""
    if not pub_href:
        return "(No abstract found)"
    
    abstract_text = "(No abstract found)"
    
    try:
        # Open in new page
        new_page = await browser.new_page()
        full_url = f"https://scholar.google.com{pub_href}" if pub_href.startswith("/") else pub_href
        
        await new_page.goto(full_url, wait_until="domcontentloaded", timeout=10000)
        await asyncio.sleep(1)
        
        # Scroll to trigger lazy loading
        await new_page.evaluate("window.scrollTo(0, 300)")
        await asyncio.sleep(1)
        
        # Try multiple selectors
        selectors = [".gsh_csp", ".gsh_csp_ab", ".gsh_small", "#gsc_oci_value", ".gsc_oci_value"]
        
        for selector in selectors:
            try:
                elem = new_page.locator(selector).first
                if await elem.count() > 0:
                    text = await elem.inner_text()
                    if text and len(text.strip()) > 10:
                        abstract_text = text.strip()
                        break
            except:
                continue
        
        await new_page.close()
        
    except Exception as e:
        print(f"    ‚ö†Ô∏è  Could not fetch abstract: {str(e)[:50]}")
    
    return abstract_text


def enrich_with_scopus_data(publications, scopus_author_id):
    """Enrich publications with Scopus data"""
    for i, pub in enumerate(publications, 1):
        print(f"  {i}/{len(publications)} Checking Scopus for: {pub['title'][:50]}...")
        
        scopus_data = get_scopus_publication_details(scopus_author_id, pub["title"])
        
        if scopus_data:
            pub["scopus_id"] = scopus_data.get("scopus_id", "N/A")
            pub["scopus_eid"] = scopus_data.get("eid", "N/A")
            pub["scopus_doi"] = scopus_data.get("doi", "N/A")
            pub["scopus_year"] = scopus_data.get("publication_year", "N/A")
            print(f"    ‚úÖ Found in Scopus (Year: {pub['scopus_year']})")
        else:
            print(f"    ‚ö†Ô∏è  Not found in Scopus")
    
    return publications


def save_to_excel(publications, author_name, scopus_author_id):
    """Save publications to Excel file with formatting"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Publications"
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Headers
    headers = [
        "No.",
        "Author Name",
        "Scopus Author ID",
        "Publication Title",
        "Abstract",
        "Publication Year (Scholar)",
        "Publication Year (Scopus)",
        "Citations (Scholar)",
        "Scopus Document ID",
        "Scopus EID",
        "DOI"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 80
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 30
    ws.column_dimensions['K'].width = 25
    
    # Add data
    for i, pub in enumerate(publications, 1):
        row = i + 1
        ws.cell(row, 1, i)
        ws.cell(row, 2, author_name)
        ws.cell(row, 3, scopus_author_id or "N/A")
        ws.cell(row, 4, pub["title"])
        ws.cell(row, 5, pub["abstract"])
        ws.cell(row, 6, pub["year"])
        ws.cell(row, 7, pub.get("scopus_year", "N/A"))
        ws.cell(row, 8, pub["citations"])
        ws.cell(row, 9, pub.get("scopus_id", "N/A"))
        ws.cell(row, 10, pub.get("scopus_eid", "N/A"))
        ws.cell(row, 11, pub.get("scopus_doi", "N/A"))
        
        # Apply alignment
        for col in range(1, 12):
            ws.cell(row, col).alignment = Alignment(vertical="top", wrap_text=True)
    
    # Save file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"publications_{author_name.replace(' ', '_')}_{timestamp}.xlsx"
    wb.save(filename)
    
    return filename


# ---------------- Main Entry Point ----------------
async def main():
    parser = argparse.ArgumentParser(
        description="Scrape publications from Google Scholar and enrich with Scopus data"
    )
    parser.add_argument(
        "author_name",
        type=str,
        nargs="?",
        help="Author name to search for (e.g., 'John Smith')"
    )
    
    args = parser.parse_args()
    
    # If no author name provided, use default or prompt
    if not args.author_name:
        author_name = input("Enter author name: ").strip()
        if not author_name:
            raise SystemExit("‚ùå Author name is required.")
    else:
        author_name = args.author_name
    
    await scrape_google_scholar_playwright(author_name)


if __name__ == "__main__":
    asyncio.run(main())
